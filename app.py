import os
import json
from datetime import datetime
from typing import Dict, List

from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session, jsonify
from werkzeug.utils import secure_filename

from utils.db import (
    initialize_database,
    get_connection,
    upsert_product,
    fetch_all_products,
    fetch_product_by_product_id,
    delete_product_by_product_id,
    create_sale_with_items,
    fetch_sales_report,
    decrement_product_stock,
)
from utils.printer import PrinterClient, build_receipt_lines

from dotenv import load_dotenv

# Load environment variables
load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
QR_DIR = os.path.join(BASE_DIR, "static", "qr")
STATIC_DIR = os.path.join(BASE_DIR, "static")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(QR_DIR, exist_ok=True)

app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key")

# Market info configuration
MARKET_NAME = os.environ.get("MARKET_NAME", "سوق العاصمة")
MARKET_ADDRESS = os.environ.get("MARKET_ADDRESS", "القاهرة، مصر")
MARKET_PHONE = os.environ.get("MARKET_PHONE", "+20 100 000 0000")

# Printer configuration
PRINTER_BACKEND = os.environ.get("PRINTER_BACKEND", "escpos_usb")  # escpos_usb, escpos_network, cups
PRINTER_ARABIC = os.environ.get("PRINTER_ARABIC", "1") == "1"

# Initialize DB
initialize_database()


@app.context_processor
def inject_globals():
    return {
        "market_name": MARKET_NAME,
    }


def _ensure_cart() -> Dict[str, int]:
    if "cart" not in session:
        session["cart"] = {}
    return session["cart"]


def _save_cart(cart: Dict[str, int]) -> None:
    session["cart"] = cart
    session.modified = True


@app.route("/")
def index():
    return redirect(url_for("sales"))


@app.route("/static/qr/<path:filename>")
def serve_qr(filename: str):
    return send_from_directory(QR_DIR, filename)


@app.route("/products")
def products_list():
    products = fetch_all_products()
    return render_template(
        "products.html",
        title="إدارة المنتجات",
        products=products,
    )


@app.route("/products/add", methods=["GET", "POST"])
def product_add():
    if request.method == "POST":
        product_id = request.form.get("product_id", "").strip()
        name = request.form.get("name", "").strip()
        price = request.form.get("price", "0").strip()
        quantity = request.form.get("quantity", "0").strip()
        date_added = request.form.get("date_added") or datetime.now().strftime("%Y-%m-%d")

        try:
            price_value = float(price)
            quantity_value = int(quantity)
        except ValueError:
            flash("برجاء إدخال سعر وكمية صحيحة.", "danger")
            return redirect(url_for("product_add"))

        try:
            upsert_product(product_id, name, price_value, quantity_value, date_added, qr_dir=QR_DIR)
            flash("تم حفظ المنتج وإنشاء QR بنجاح.", "success")
        except Exception as exc:
            flash(f"حدث خطأ أثناء الحفظ: {exc}", "danger")
        return redirect(url_for("products_list"))

    return render_template(
        "product_form.html",
        title="إضافة منتج",
        action_label="حفظ",
        product=None,
    )


@app.route("/products/<product_id>/edit", methods=["GET", "POST"])
def product_edit(product_id: str):
    product = fetch_product_by_product_id(product_id)
    if product is None:
        flash("المنتج غير موجود.", "warning")
        return redirect(url_for("products_list"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        price = request.form.get("price", "0").strip()
        quantity = request.form.get("quantity", "0").strip()
        date_added = request.form.get("date_added") or product["date_added"]

        try:
            price_value = float(price)
            quantity_value = int(quantity)
        except ValueError:
            flash("برجاء إدخال سعر وكمية صحيحة.", "danger")
            return redirect(url_for("product_edit", product_id=product_id))

        try:
            upsert_product(product_id, name, price_value, quantity_value, date_added, qr_dir=QR_DIR)
            flash("تم تحديث المنتج.", "success")
        except Exception as exc:
            flash(f"حدث خطأ أثناء التحديث: {exc}", "danger")
        return redirect(url_for("products_list"))

    return render_template(
        "product_form.html",
        title="تعديل منتج",
        action_label="تحديث",
        product=product,
    )


@app.route("/products/<product_id>/delete", methods=["POST"])
def product_delete(product_id: str):
    try:
        delete_product_by_product_id(product_id)
        flash("تم حذف المنتج.", "success")
    except Exception as exc:
        flash(f"تعذر حذف المنتج: {exc}", "danger")
    return redirect(url_for("products_list"))


@app.route("/import", methods=["POST"])
def import_products():
    """Import products from uploaded Excel file using openpyxl."""
    file = request.files.get("excel_file")
    if not file:
        flash("برجاء اختيار ملف Excel.", "warning")
        return redirect(url_for("products_list"))

    filename = secure_filename(file.filename or "products.xlsx")
    path = os.path.join(UPLOADS_DIR, filename)
    file.save(path)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # Map headers
        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            headers[cell.value] = idx
        required = ["Product ID", "Product Name", "Price (EGP)", "Quantity", "Date Added"]
        missing = [h for h in required if h not in headers]
        if missing:
            raise ValueError(f"الأعمدة المفقودة: {', '.join(missing)}")

        count = 0
        for row_idx in range(2, ws.max_row + 1):
            pid = str(ws.cell(row=row_idx, column=headers["Product ID"]).value or "").strip()
            if not pid:
                continue
            name = str(ws.cell(row=row_idx, column=headers["Product Name"]).value or "").strip()
            price_cell = ws.cell(row=row_idx, column=headers["Price (EGP)"]).value
            qty_cell = ws.cell(row=row_idx, column=headers["Quantity"]).value
            date_cell = ws.cell(row=row_idx, column=headers["Date Added"]).value

            try:
                price = float(price_cell)
            except Exception:
                price = 0.0
            try:
                qty = int(qty_cell)
            except Exception:
                qty = 0
            if hasattr(date_cell, "strftime"):
                date_added = date_cell.strftime("%Y-%m-%d")
            else:
                date_added = datetime.now().strftime("%Y-%m-%d")

            upsert_product(pid, name, price, qty, date_added, qr_dir=QR_DIR)
            count += 1

        flash(f"تم استيراد {count} منتج(ات) وتوليد QR.", "success")
    except Exception as exc:
        flash(f"فشل الاستيراد: {exc}", "danger")

    return redirect(url_for("products_list"))


@app.route("/sales")
def sales():
    cart = _ensure_cart()
    # Build items with product info
    cart_items: List[Dict] = []
    subtotal: float = 0.0
    for pid, qty in cart.items():
        product = fetch_product_by_product_id(pid)
        if product:
            line_total = product["price_egp"] * qty
            subtotal += line_total
            cart_items.append({
                "product_id": pid,
                "name": product["name"],
                "price_egp": product["price_egp"],
                "quantity": qty,
                "line_total": line_total,
            })

    return render_template(
        "sales.html",
        title="صفحة البيع",
        cart_items=cart_items,
        subtotal=subtotal,
    )


@app.route("/cart/add", methods=["POST"])
def cart_add():
    data = request.get_json(silent=True) or request.form
    product_id = (data.get("product_id") or "").strip()
    quantity = int(data.get("quantity", 1))

    product = fetch_product_by_product_id(product_id)
    if not product:
        return jsonify({"ok": False, "message": "المنتج غير موجود"}), 404

    cart = _ensure_cart()
    cart[product_id] = cart.get(product_id, 0) + max(1, quantity)
    _save_cart(cart)
    return jsonify({"ok": True})


@app.route("/cart/update", methods=["POST"])
def cart_update():
    data = request.get_json(silent=True) or request.form
    product_id = (data.get("product_id") or "").strip()
    quantity = max(0, int(data.get("quantity", 0)))
    cart = _ensure_cart()
    if quantity == 0:
        cart.pop(product_id, None)
    else:
        cart[product_id] = quantity
    _save_cart(cart)
    return jsonify({"ok": True})


@app.route("/cart/clear", methods=["POST"])
def cart_clear():
    _save_cart({})
    return jsonify({"ok": True})


@app.route("/sale/complete", methods=["POST"])
def sale_complete():
    cart = _ensure_cart()
    if not cart:
        return jsonify({"ok": False, "message": "لا توجد عناصر في السلة"}), 400

    # Build sale items and totals
    items: List[Dict] = []
    total: float = 0.0
    for pid, qty in cart.items():
        product = fetch_product_by_product_id(pid)
        if not product:
            continue
        line_total = product["price_egp"] * qty
        total += line_total
        items.append({
            "product_id": pid,
            "name": product["name"],
            "price_egp": product["price_egp"],
            "quantity": qty,
            "line_total": line_total,
        })

    if not items:
        return jsonify({"ok": False, "message": "لا توجد عناصر صالحة"}), 400

    # Create sale in DB
    sale_id = create_sale_with_items(
        sale_date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        items=items,
        total_egp=total,
    )

    # Decrement stock
    for item in items:
        decrement_product_stock(item["product_id"], item["quantity"])

    # Print receipt
    printer = PrinterClient(
        backend=PRINTER_BACKEND,
        prefer_arabic=PRINTER_ARABIC,
    )
    receipt_lines_ar, receipt_lines_en = build_receipt_lines(
        market_name=MARKET_NAME,
        market_address=MARKET_ADDRESS,
        market_phone=MARKET_PHONE,
        sale_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
        items=items,
        total_egp=total,
    )

    used_lang = printer.print_receipt(receipt_lines_ar=receipt_lines_ar, receipt_lines_en=receipt_lines_en)

    # Update print language stored in sale record (simple approach: update field)
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("UPDATE sales SET print_lang = ? WHERE id = ?", (used_lang, sale_id))
        conn.commit()
    finally:
        conn.close()

    # Clear cart
    _save_cart({})

    return jsonify({"ok": True, "sale_id": sale_id, "used_lang": used_lang})


@app.route("/reports")
def reports():
    period = request.args.get("period", "daily")  # daily or monthly
    date_str = request.args.get("date")

    if not date_str:
        if period == "monthly":
            date_str = datetime.now().strftime("%Y-%m")
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")

    report_rows, total_sum = fetch_sales_report(period=period, date_str=date_str)

    return render_template(
        "reports.html",
        title="تقارير المبيعات",
        period=period,
        date_str=date_str,
        report_rows=report_rows,
        total_sum=total_sum,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    debug_flag = os.environ.get("FLASK_DEBUG", "0") == "1"
    use_reloader = os.environ.get("FLASK_USE_RELOADER", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug_flag, use_reloader=use_reloader)
